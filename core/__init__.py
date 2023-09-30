
# CORE Package PySide2
# ///////////////////////////////////////////////////////////////
from PySide2.QtCore import QMutex,QWaitCondition,QMetaObject,QEventLoop,QEvent, QCoreApplication, QPropertyAnimation, QDate, Qt, Signal, QThread, Slot, QObject, QAbstractItemModel, QAbstractTableModel, QModelIndex, Signal, QTimer,QRunnable,QThreadPool
from PySide2.QtWidgets import QPushButton,QStyle, QWidget, QPlainTextEdit, QStyledItemDelegate, QApplication, QToolTip, QTableView, QDialog, QMainWindow, QMessageBox, QFileDialog, QTableWidgetItem, QHeaderView, QAbstractItemView, QTableWidget, QComboBox, QSpinBox, QCheckBox, QMessageBox, QFrame, QAbstractSpinBox
from PySide2.QtGui import QIcon,QStandardItemModel, QStandardItem, QTextCursor, QColor,QPen
# Database SQLite3
# ///////////////////////////////////////////////////////////////
import sqlite3
import os
from sqlite3 import Error
# Libs Python
# ///////////////////////////////////////////////////////////////
import re
import csv
import json
import xlrd
import shutil
import pandas as pd
import openpyxl as xl
from io import StringIO
from xlsx2csv import Xlsx2csv
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from pprint import pprint,PrettyPrinter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# System OS 
# ///////////////////////////////////////////////////////////////
import os
import sys
import uuid
import win32com.client as win32
from time import sleep
from datetime import datetime
# CORE Package UI gerado pelo Qt Designer
# ///////////////////////////////////////////////////////////////
from ui.MainWindow import Ui_mainWindow